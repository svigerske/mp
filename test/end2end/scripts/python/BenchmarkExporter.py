from Exporter import Exporter
import ModelRunner
import openpyxl
import os
import math
import Solver
class BenchmarkExporter(Exporter):

    def get_last_progress(self) -> list:
        if self._fileName is None:
            return None
        if not os.path.exists(self.get_file_name()):
            return None
        
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook(self.get_file_name())
        sheet = workbook.active
        first_column_values = [cell.value for cell in sheet['A'][1:]]
        workbook.close()
        return first_column_values
     
    def _initialize_from_file(self) -> bool:
        if os.path.exists(self.get_file_name()):
            self.workbook = openpyxl.load_workbook(self.get_file_name())
            self.sheet_main = self.workbook.get_sheet_by_name("Sheet")
            self.sheet_stats = self.workbook.get_sheet_by_name("stats")
                
            # Iterate through rows to find the first empty row
            first_column_values = [cell.value for cell in self.sheet_main['A'][1:]]
            self.current_row=len(first_column_values)
            if self.current_row >0:
                self.current_row+=1

            headers = [cell.value for cell in self.sheet_stats[1]]
        
            # Iterate through the rows, starting from the second row
            for row in self.sheet_stats.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    # Stop if the first cell in the row is empty (end of data)
                    break
                solver = row[0]
                self.solvers[solver] = dict(zip(headers[1:], row[1:]))
            
            self.bold_style = "bold_style"
            return True
        return False
            
    def __init__(self, fileName=""):
        super().__init__()
        self._fileName=fileName
        self.solvers={}
        if not self._initialize_from_file():
            self.workbook = openpyxl.Workbook()
            self.sheet_main = self.workbook.active
            self.sheet_stats = self.workbook.create_sheet("stats")
            self.current_row = 0
            self.bold_style = openpyxl.styles.NamedStyle(name="bold_style")
            self.bold_style.font = openpyxl.styles.Font(bold=True)
        self.style_neutral = "Neutral"
        self.style_bad =     "Bad"

    def get_file_name(self):
        base_name, _= os.path.splitext(self._fileName)
        if os.path.isabs(base_name):
            return base_name + '.xlsx'
        return os.path.abspath(base_name) + '.xlsx'

    def sanifyString(self, s):
        try:
            s = s.replace("\n", " - ")
            s = s.replace(",", ";")
            return s
        except:
          return s

    def writeHeader(self, mr: ModelRunner):
        hdr = "Name,Expected_Obj,Variables,Int_Variables,Constraints,Nnz"
        
        runs=mr.getRuns()
        
        for runner, run in zip(mr.getRunners(), runs):
            sname = run[-1]["solver"]
            hdr += f",{sname}-Obj,{sname}-Time,{sname}-Status"
            if runner.support_times():
                hdr += f",{sname}-SetupTime"
            
        for run in runs:
            hdr += ",{}-SolverMsg".format(run[-1]["solver"])

        header_list = hdr.split(',')
        for col_num, header_text in enumerate(header_list, 1):
            cell=self.sheet_main.cell(row=1, column=col_num, value=header_text)
            cell.style =self.bold_style
        self.current_row += 1

    def getModelsStats(self, run):
        if not "modelStats" in run[-1]:
            return None
        stats = run[-1]["modelStats"]
        return [ stats["nvars"], stats["nintvars"], stats["nconstr"], stats["nnz"]]
  
    def getStyle(self, r):
        try:
            res = r[-1]["parsed_result"]
            if res == "OK":
                if "solved" in r[-1]["timelimit"]:
                    return None
                elif "limit" in r[-1]["timelimit"]:
                    return self.style_neutral
            else:
                return self.style_bad
        except: 
            return self.style_bad

    def writeLastResultLine(self, mr: ModelRunner):
        i = len( mr.getRuns()[0] )
        m = mr.getModels()[i-1]
        res = [m.getName(), m.getExpectedObjective()]
        stats = self.getModelsStats(mr.getRuns()[0])
        if stats != None:
            res.extend(stats)
        else:
            res.extend(["-","-","-","-"])  
            
        styles = [None for _ in res]

        for runner, r in zip(mr.getRunners(), mr.getRuns()):
            style=self.getStyle(r)

            rr = r[-1]["parsed_result"]
            if rr=="OK":
                rr = self._getDictMemberOrMissingStr(r[-1], "timelimit")
            res.extend([
              self._getDictMemberOrMissingStr(r[-1], "objective"),
              self._getDictMemberOrMissingStr(r[-1], "solutionTime"),
              rr
              ])
            styles.extend([style for _ in range(3)])
            if runner.support_times():
                if "times" in r[-1]:
                    res.append(r[-1]["times"].get("setup",0))
                else:
                    res.append(0)
                styles.append(style)

           
        for r in mr.getRuns():
             res.append(
                self._getDictMemberOrMissingStr(r[-1], "outmsg"))
             styles.append(self.getStyle(r))

        for col_num, header_text in enumerate(res,1):
            cell=self.sheet_main.cell(row=self.current_row, column=col_num, 
                                 value=header_text)
            if styles[col_num-1] is not None:
                cell.style=styles[col_num-1]
        self.current_row += 1
                
    def _getDictMemberOrMissingStr(self, dct, key):
        try:
            return dct[key]
        except:
            return "-"
    def addToDict(self, sname, item, value):
        self.solvers[sname][item]=self.solvers[sname][item]+value

    def collectSolverStats(self, mr):
        runs = mr.getRuns()
        i = len( runs[0] ) 
        m = mr.getModels()[i-1]
        exp_obj= m.getExpectedObjective()
        
        for r in runs:
            lastRun = r[-1]
            sname=lastRun["solver"]
            if isinstance(sname, Solver.Solver):
                sname=sname.getName()
            if "Skipped" in lastRun["outmsg"]:
                self.addToDict(sname, "failed", 1)
                return
            
            if "script failure" in lastRun["outmsg"]:
                self.addToDict(sname, "failed_ampl", 1)
                return
            outcome=lastRun["timelimit"]
            self.addToDict(sname, "time_all", lastRun["solutionTime"]) 
           
            if isinstance(r, str):
                self.addToDict(sname, "failed", 1)
            else:
                if "solved" in outcome:
                   self.addToDict(sname, "solved", 1)
                   self.addToDict(sname, "time_solved", lastRun["solutionTime"]) 
                elif "limit" in outcome:
                    try:
                        if  math.isclose(exp_obj,lastRun["objective"], rel_tol=0.001):
                            self.addToDict(sname, "timelimit_correct", 1)
                        else:
                            self.addToDict(sname, "timelimit", 1)
                    except:
                        self.addToDict(sname, "timelimit", 1)
                else:
                    self.addToDict(sname, "failed", 1)

    def initialize(self, mr: ModelRunner):
         runs=mr.getRuns()
         runners = mr.getRunners()
         solvers_runs = zip(runners, runs)

         for (runner, solver_run) in solvers_runs:
            last_run = solver_run[-1]
            solver=last_run["solver"]
            if isinstance(solver, Solver.Solver):
                solvername=solver.getName()
            else:
                solvername=solver
                solver=runner
            if len(self.solvers)==0:
                self.solvers[solvername]={
                    "solved" : 0,
                    "failed" : 0,
                    "failed_ampl" : 0,
                    "timelimit" : 0,
                    "timelimit_correct" : 0,
                    "time_solved" : 0,
                    "time_all" : 0
                }
                
            if self.collector_:
                self.collector_.add_solver_def(solvername, solver.get_version())
             
    def writeSolverStats(self):
        header = ["Solver"]
        header.extend(self.solvers[list(self.solvers.keys())[0]].keys())

        for col_num, header_text in enumerate(header, 1):
            cell=self.sheet_stats.cell(row=1, column=col_num, value=header_text)
            cell.style =self.bold_style 
        for row_num, (name, properties) in enumerate(self.solvers.items(), 2):
            self.sheet_stats.cell(row=row_num, column=1, value=name)
            for col_num, value in enumerate(properties.values(), 2):
                self.sheet_stats.cell(row=row_num, column=col_num, value=value)

    def _exportInstanceResults(self, mr: ModelRunner):
        i = len( mr.getRuns()[0] )
        
        if i == 1 and self.current_row==0:
            self.writeHeader(mr) 
            
        self.initialize(mr)
        self.writeLastResultLine(mr)
        self.collectSolverStats(mr)
        self.writeSolverStats()
        self.workbook.save(self.get_file_name())

                   



